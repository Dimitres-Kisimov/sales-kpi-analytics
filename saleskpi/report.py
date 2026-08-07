"""report.py — turn the analytics into DELIVERABLES a management team receives.

`analyze()` runs the whole pipeline (KPIs, mix, ABC-XYZ, RFM, margin bridge,
per-series forecast selection, reorder recommendations, auto decision cards).
`write_deliverables()` emits:

    deliverables/management_report.md   a Quarterly-Business-Review write-up
    deliverables/kpi_workbook.xlsx      multi-sheet Excel (openpyxl, optional)
    deliverables/reorder_list.csv       below-ROP SKUs to order now
    deliverables/analysis.json          machine-readable everything
    deliverables/forecast.png           revenue history + forecast (matplotlib, optional)

Author: Dimitres Kisimov.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from . import anomaly, metrics
from .dataset import load, monthly_series
from .forecast import select_and_forecast
from .inventory import reorder_recommendation
from .pvm import bridge_waterfall_steps, plain_language, revenue_bridge
from .spend import spend_summary, waterfall_steps

OUT = Path(__file__).resolve().parents[1] / "deliverables"
LEAD_MONTHS = 1.0          # modeling assumption: ~1 month supplier replenishment lead


def analyze(rows: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    rows = rows if rows is not None else load()
    kpi = metrics.kpi_summary(rows)
    grow = metrics.growth(rows)
    regions = metrics.by_dimension(rows, "region")
    categories = metrics.by_dimension(rows, "product_category")
    channels = metrics.by_dimension(rows, "channel")
    reps = metrics.by_dimension(rows, "sales_rep")
    abcxyz = metrics.abc_xyz(rows, "product_category")
    abc_by_cat = {a["product_category"]: a for a in abcxyz}
    rfm_all = metrics.rfm(rows)
    seg_counts: dict[str, int] = {}
    for c in rfm_all:
        seg_counts[c["segment"]] = seg_counts.get(c["segment"], 0) + 1
    bridge = metrics.margin_bridge(rows)
    rev_bridge = revenue_bridge(rows)
    conc = metrics.concentration(rows)
    expenditure = spend_summary(rows)
    kpi_alerts = anomaly.kpi_exceptions(rows)

    # headline revenue forecast (total monthly revenue) with CV model selection
    rev_series = [v for _, v in monthly_series(rows, "revenue_eur")]
    rev_fc = select_and_forecast(rev_series, horizon=3)

    # per (region x category) unit-demand forecast + reorder decision
    region_names = [r["region"] for r in regions]
    cat_names = [c["product_category"] for c in categories]
    forecasts, reorders = [], []
    for reg in region_names:
        for cat in cat_names:
            sub = [r for r in rows if r["region"] == reg and r["product_category"] == cat]
            if len(sub) < 6:
                continue
            units = [v for _, v in monthly_series(sub, "units")]
            sel = select_and_forecast(units, horizon=1)
            abc = abc_by_cat.get(cat, {}).get("abc", "B")
            on_hand = round(sum(units[-3:]) / 3 * (1.3 if abc == "A" else 0.7), 1)
            rec = reorder_recommendation(units, on_hand, LEAD_MONTHS, abc_class=abc)
            forecasts.append({"region": reg, "category": cat, "winner": sel.winner,
                              "cv_mase": sel.cv_mase, "pattern": sel.demand["pattern"],
                              "next_units": sel.values[0]})
            if rec.get("reorder"):
                reorders.append({"region": reg, "category": cat, **rec})
    reorders.sort(key=lambda d: -d["recommended_order_qty"])

    cards = _decision_cards(kpi, grow, bridge, abcxyz, seg_counts, regions, conc, rev_fc,
                            expenditure, rev_bridge, kpi_alerts)
    return {
        "kpi": kpi, "growth": grow, "regions": regions, "categories": categories,
        "channels": channels, "reps": reps, "abc_xyz": abcxyz,
        "rfm_segments": seg_counts, "rfm_top": rfm_all[:10],
        "margin_bridge": bridge, "revenue_bridge": rev_bridge,
        "concentration": conc, "expenditure": expenditure, "kpi_alerts": kpi_alerts,
        "revenue_forecast": {"winner": rev_fc.winner, "cv_mase": rev_fc.cv_mase,
                             "history": rev_series, "forecast": rev_fc.values,
                             "leaderboard": [vars(r) for r in rev_fc.leaderboard]},
        "series_forecasts": forecasts, "reorder_list": reorders,
        "decision_cards": cards,
    }


def _decision_cards(kpi, grow, bridge, abcxyz, seg_counts, regions, conc, rev_fc,
                    expenditure=None, rev_bridge=None, kpi_alerts=None) -> list[str]:
    cards = []
    if kpi_alerts and kpi_alerts["summary"]["total_alerts"]:
        s = kpi_alerts["summary"]
        worst = kpi_alerts["alerts"][0]
        hit = ", ".join(sorted({a["label"] for a in kpi_alerts["alerts"]}))
        wv = (f"{worst['value']:,.0f} EUR" if worst["unit"] == "EUR"
              else f"{worst['value']:.2f}%")
        cards.append(
            f"KPI exception monitor: {s['total_alerts']} out-of-control month(s) "
            f"({s['unfavorable']} unfavourable) on {hit} vs robust "
            f"{kpi_alerts['k_sigma']:.1f}-sigma control limits - worst {worst['label']} "
            f"{wv} at {worst['month']} (modified z {worst['modified_z']:+.1f}); "
            f"{len(s['kpis_in_control'])} monitored KPIs in control.")
    if rev_bridge and rev_bridge.get("available"):
        drivers = {"price": rev_bridge["price_effect_eur"],
                   "volume": rev_bridge["volume_effect_eur"],
                   "mix": rev_bridge["mix_effect_eur"]}
        best = max(drivers, key=drivers.get)
        worst = min(drivers, key=drivers.get)
        cards.append("Revenue bridge (YoY): " + plain_language(rev_bridge)
                     + f" {best.capitalize()} was the biggest tailwind "
                     f"({drivers[best]:+,.0f} EUR); {worst} the biggest drag "
                     f"({drivers[worst]:+,.0f} EUR).")
    dd = (expenditure or {}).get("leakage_drilldown")
    if dd and dd.get("by_sales_rep"):
        w = dd["by_sales_rep"][0]
        cards.append(f"Discount drill-down: {w['sales_rep']} ({w['region']}) leads the leakage "
                     f"table - {w['leakage_eur']:,.0f} EUR vs list, {w['excess_eur']:,.0f} EUR of "
                     f"it above the {dd['policy_discount_pct']:.0f}% policy assumption - review "
                     f"discount authority there first.")
    yoy = grow.get("yoy_pct")
    if yoy is not None:
        cards.append(f"Revenue is {'up' if yoy >= 0 else 'down'} {abs(yoy):.1f}% YoY; "
                     f"gross margin sits at {kpi['gross_margin_pct']:.1f}%.")
    if bridge.get("available"):
        drivers = {"price": bridge["price_effect_eur"], "volume": bridge["volume_effect_eur"],
                   "mix": bridge["mix_effect_eur"]}
        worst = min(drivers, key=drivers.get)
        best = max(drivers, key=drivers.get)
        cards.append(f"Margin bridge: {best} added {drivers[best]:,.0f} EUR while {worst} "
                     f"cost {abs(drivers[worst]):,.0f} EUR - steer commercial focus accordingly.")
    cz = [a["product_category"] for a in abcxyz if a["class"] in ("CZ", "CY", "BZ")]
    if cz:
        cards.append(f"Long-tail/erratic categories ({', '.join(cz)}) tie up working "
                     f"capital - move to min-stock or make-to-order.")
    at_risk = seg_counts.get("At risk", 0) + seg_counts.get("Churned / dormant", 0)
    cards.append(f"{at_risk} customers are At-risk/Churned (of {sum(seg_counts.values())}); "
                 f"hand the named list to sales for win-back.")
    cards.append(f"Top-20% of customers drive {conc['top_share_pct']:.0f}% of revenue - "
                 f"concentration risk to monitor.")
    cards.append(f"Next-quarter revenue forecast (model: {rev_fc.winner}, CV-MASE "
                 f"{rev_fc.cv_mase:.2f}): {', '.join(f'{v:,.0f}' for v in rev_fc.values)} EUR.")
    if kpi["otif_pct"] < 90:
        cards.append(f"OTIF is {kpi['otif_pct']:.0f}% (below the 90% service bar) - "
                     f"investigate peak-month delivery reliability.")
    return cards


# --------------------------------------------------------------------------- #
# Deliverable writers
# --------------------------------------------------------------------------- #
def write_deliverables(analysis: dict[str, Any], outdir: Path | None = None) -> list[str]:
    out = outdir or OUT
    out.mkdir(parents=True, exist_ok=True)
    made = []
    (out / "analysis.json").write_text(json.dumps(analysis, indent=2), encoding="utf-8")
    made.append("analysis.json")
    made.append(_write_reorder_csv(analysis, out))
    made.append(_write_markdown(analysis, out))
    x = _write_xlsx(analysis, out)
    if x:
        made.append(x)
    p = _write_chart(analysis, out)
    if p:
        made.append(p)
    lsvg = _write_leakage_svg(analysis, out)
    if lsvg:
        made.append(lsvg)
    lw = _write_leakage_chart(analysis, out)
    if lw:
        made.append(lw)
    made.append(_write_pvm_csv(analysis, out))
    psvg = _write_pvm_svg(analysis, out)
    if psvg:
        made.append(psvg)
    ppng = _write_pvm_chart(analysis, out)
    if ppng:
        made.append(ppng)
    csvg = _write_control_chart_svg(analysis, out)
    if csvg:
        made.append(csvg)
    cpng = _write_control_chart_chart(analysis, out)
    if cpng:
        made.append(cpng)
    return made


def _write_reorder_csv(a: dict, out: Path) -> str:
    rows = a["reorder_list"]
    cols = ["region", "category", "abc_class", "service_level", "avg_demand",
            "safety_stock", "reorder_point", "on_hand", "recommended_order_qty"]
    with (out / "reorder_list.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    return "reorder_list.csv"


def _write_markdown(a: dict, out: Path) -> str:
    k, g = a["kpi"], a["growth"]
    b = a["margin_bridge"]
    rb = a.get("revenue_bridge", {})
    lines = [
        "# Quarterly Business Review — Sales & Demand Analytics",
        "",
        "*Auto-generated from 24 months of order data. Author: Dimitres Kisimov.*",
        "",
        "## 1. Headline KPIs",
        f"- Revenue: **{k['revenue_eur']:,.0f} EUR** "
        + (f"({g['yoy_pct']:+.1f}% YoY)" if g.get("yoy_pct") is not None else ""),
        f"- Gross margin: **{k['gross_margin_pct']:.1f}%** ({k['gross_margin_eur']:,.0f} EUR)",
        f"- Orders: {k['orders']:,} · AOV: {k['aov_eur']:,.0f} EUR · Active customers: {k['active_customers']}",
        f"- Service: OTIF {k['otif_pct']:.0f}% · on-time {k['on_time_pct']:.0f}% · "
        f"fill rate {k['fill_rate_pct']:.0f}% · returns {k['returns_rate_pct']:.1f}%",
        f"- Discount leakage: {k['discount_leakage_pct']:.1f}% of list value",
        "",
        "## 2. What the numbers say (decision cards)",
    ]
    lines += [f"- {c}" for c in a["decision_cards"]]
    lines += ["", "## 3. Revenue bridge — why did revenue move? (YoY, price / volume / mix)"]
    if rb.get("available"):
        lines += [
            f"- {plain_language(rb)}",
            f"- Prior 12m revenue: {rb['prior_revenue_eur']:,.0f} EUR ({rb['period_prior']}) → "
            f"current {rb['current_revenue_eur']:,.0f} EUR ({rb['period_current']}) "
            f"(**{rb['total_change_eur']:+,.0f} EUR**, {rb['total_change_pct']:+.1f}%; units "
            f"{rb['volume_growth_pct']:+.1f}%)",
            f"- Price effect: {rb['price_effect_eur']:+,.0f} EUR · Volume: "
            f"{rb['volume_effect_eur']:+,.0f} EUR · Mix: {rb['mix_effect_eur']:+,.0f} EUR "
            "— the three sum to the total change by construction (mix is the residual).",
            "",
            "| Category | Prior rev | Current rev | Price | Volume | Mix |",
            "|---|--:|--:|--:|--:|--:|",
        ]
        for c in rb["by_category"]:
            lines.append(f"| {c['product_category']} | {c['prior_revenue_eur']:,.0f} | "
                         f"{c['current_revenue_eur']:,.0f} | {c['price_effect_eur']:+,.0f} | "
                         f"{c['volume_effect_eur']:+,.0f} | {c['mix_effect_eur']:+,.0f} |")
        lines.append("")
        lines.append("*Realised price = revenue ÷ units per category; volume is proportional "
                     "growth at the prior blended price, mix the reallocation across categories. "
                     "See `pvm_bridge.csv` and `pvm_waterfall.svg`.*")
    lines += ["", "## 4. Margin bridge (YoY, price / volume / mix)"]
    if b.get("available"):
        lines += [
            f"- Prior 12m margin: {b['prior_margin_eur']:,.0f} EUR → current {b['current_margin_eur']:,.0f} EUR "
            f"(**{b['total_change_eur']:+,.0f} EUR**)",
            f"- Price effect: {b['price_effect_eur']:+,.0f} EUR · Volume: {b['volume_effect_eur']:+,.0f} EUR · "
            f"Mix: {b['mix_effect_eur']:+,.0f} EUR",
        ]
    lines += ["", "## 5. Portfolio (ABC × XYZ)",
              "| Category | Revenue | Share | Class | CV |", "|---|--:|--:|:--:|--:|"]
    for r in a["abc_xyz"]:
        lines.append(f"| {r['product_category']} | {r['revenue_eur']:,.0f} | "
                     f"{r['revenue_share_pct']:.1f}% | {r['class']} | {r['cv']} |")
    rf = a["revenue_forecast"]
    lines += ["", "## 6. Revenue forecast (next 3 months)",
              f"- Model selected by rolling-origin cross-validation: **{rf['winner']}** "
              f"(CV-MASE {rf['cv_mase']:.2f}; <1 beats the seasonal-naive baseline)",
              f"- Forecast: {', '.join(f'{v:,.0f}' for v in rf['forecast'])} EUR",
              "", "## 7. Replenishment — reorder now",
              f"- {len(a['reorder_list'])} region×category cells are at/below their reorder point.",
              "  See `reorder_list.csv` for quantities (safety stock at per-ABC service levels)."]
    sp = a.get("expenditure")
    if sp:
        lines += ["", "## 8. Expenditure & spend",
                  f"- Total COGS: **{sp['total_cogs_eur']:,.0f} EUR** "
                  f"on {sp['list_value_eur']:,.0f} EUR of list value.",
                  f"- Discount leakage: **{sp['discount_leakage_eur']:,.0f} EUR** "
                  f"({sp['discount_leakage_pct']:.1f}% of list) given away versus list price.",
                  f"- Modeled returns cost: {sp['returns_cost_total_eur']:,.0f} EUR.",
                  "",
                  "| Channel | COGS | Returns cost | Discount leakage | Cost-to-serve |",
                  "|---|--:|--:|--:|--:|"]
        for c in sp["cost_to_serve_by_channel"]:
            lines.append(f"| {c['channel']} | {c['cogs_eur']:,.0f} | "
                         f"{c['returns_cost_eur']:,.0f} | {c['discount_leakage_eur']:,.0f} | "
                         f"{c['cost_to_serve_eur']:,.0f} |")
        dd = sp.get("leakage_drilldown")
        if dd:
            lines += [
                "", "### 8.1 Discount-leakage drill-down — who, where",
                f"Waterfall: gross list value {dd['gross_list_value_eur']:,.0f} EUR "
                f"− within-policy discounts {dd['within_policy_discount_eur']:,.0f} EUR "
                f"− excess discounts {dd['excess_discount_eur']:,.0f} EUR "
                f"= net revenue {dd['net_revenue_eur']:,.0f} EUR. "
                f"Both discount cuts together are the {dd['total_leakage_eur']:,.0f} EUR leakage.",
                "",
                f"*Policy threshold: {dd['policy_discount_pct']:.0f}% of list — "
                f"{dd['policy_note']}.*",
                "",
                "Top 5 reps by excess-over-policy (see `leakage_waterfall.png`):",
                "",
                "| Rep | Region | Leakage | Excess >policy | % of own revenue | "
                "Orders >policy | Median / p90 discount |",
                "|---|---|--:|--:|--:|--:|--:|",
            ]
            for r in dd["by_sales_rep"][:5]:
                lines.append(
                    f"| {r['sales_rep']} | {r['region']} | {r['leakage_eur']:,.0f} | "
                    f"{r['excess_eur']:,.0f} | {r['leakage_pct_of_revenue']:.1f}% | "
                    f"{r['orders_above_policy_pct']:.0f}% | "
                    f"{r['median_discount_pct']:.1f}% / {r['p90_discount_pct']:.1f}% |")
            lines += ["", "| Region | Leakage | Excess >policy | % of region revenue | "
                          "Orders >policy |", "|---|--:|--:|--:|--:|"]
            for r in dd["by_region"]:
                lines.append(
                    f"| {r['region']} | {r['leakage_eur']:,.0f} | {r['excess_eur']:,.0f} | "
                    f"{r['leakage_pct_of_revenue']:.1f}% | "
                    f"{r['orders_above_policy_pct']:.0f}% |")
    ex = a.get("kpi_alerts")
    if ex:
        s = ex["summary"]
        lines += ["", "## 9. KPI exceptions — statistical control monitor",
                  f"- {anomaly.plain_language(ex)}",
                  f"- Method: {ex['method']}, k = {ex['k_sigma']:.1f}; "
                  f"period {ex['period']} ({ex['months']} months).",
                  "- Scope note: the monitored KPIs are the quality/service/commercial "
                  "ratios plus AOV, where a monthly control chart is trustworthy. The "
                  "strongly-seasonal volume series (revenue, orders) are intentionally "
                  "excluded from level-anomaly detection on 24 points — their movement is "
                  "covered by the out-of-sample forecast CV and the revenue bridge."]
        if ex["alerts"]:
            lines += ["",
                      "| KPI | Month | Value | Centre | Nearest limit | Modified z | "
                      "Direction | Read | Severity |",
                      "|---|---|--:|--:|--:|--:|:--:|:--:|:--:|"]
            for al in ex["alerts"]:
                unit = "" if al["unit"] == "EUR" else al["unit"]
                lim = al["lower_limit"] if al["direction"] == "below" else al["upper_limit"]
                read = "—" if al["favorable"] is None else (
                    "favourable" if al["favorable"] else "unfavourable")
                lines.append(
                    f"| {al['label']} | {al['month']} | {al['value']:,.2f}{unit} | "
                    f"{al['center']:,.2f}{unit} | {lim:,.2f}{unit} | "
                    f"{al['modified_z']:+.2f} | {al['direction']} | {read} | "
                    f"{al['severity']} |")
        if s["kpis_in_control"]:
            lines += ["", f"*In control this period: {', '.join(s['kpis_in_control'])} "
                          f"(no month outside its robust control limits). See "
                          f"`kpi_control_chart.svg`.*"]
    (out / "management_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return "management_report.md"


def _write_xlsx(a: dict, out: Path) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F6BFF")

    ws = wb.active
    ws.title = "KPIs"
    ws.append(["Sales & Demand Analytics — KPI summary"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for key, val in a["kpi"].items():
        ws.append([key, val])

    def sheet(name, rows, cols):
        s = wb.create_sheet(name)
        s.append(cols)
        for c in s[1]:
            c.font, c.fill = hdr, fill
        for r in rows:
            s.append([r.get(c) for c in cols])

    sheet("Mix by region", a["regions"], ["region", "revenue_eur", "margin_pct", "orders"])
    sheet("ABC-XYZ", a["abc_xyz"],
          ["product_category", "revenue_eur", "revenue_share_pct", "abc", "cv", "xyz", "class"])
    sheet("RFM top", a["rfm_top"],
          ["customer_id", "recency_days", "frequency", "monetary_eur", "rfm_score", "segment"])
    sheet("Forecasts", a["series_forecasts"],
          ["region", "category", "winner", "cv_mase", "pattern", "next_units"])
    sheet("Reorder list", a["reorder_list"],
          ["region", "category", "abc_class", "safety_stock", "reorder_point",
           "on_hand", "recommended_order_qty"])
    rb = a.get("revenue_bridge", {})
    if rb.get("available"):
        sheet("Revenue bridge", rb["by_category"],
              ["product_category", "prior_units", "current_units", "prior_price_eur",
               "current_price_eur", "prior_revenue_eur", "current_revenue_eur",
               "price_effect_eur", "volume_effect_eur", "mix_effect_eur"])
        s = wb["Revenue bridge"]
        s.append([])
        s.append(["Prior revenue EUR", rb["prior_revenue_eur"]])
        s.append(["Current revenue EUR", rb["current_revenue_eur"]])
        s.append(["Total change EUR", rb["total_change_eur"]])
        s.append(["Price effect EUR", rb["price_effect_eur"]])
        s.append(["Volume effect EUR", rb["volume_effect_eur"]])
        s.append(["Mix effect EUR", rb["mix_effect_eur"]])
        s.append(["Note", "price+volume+mix = total change (mix is the residual); "
                          "synthetic data"])
    sp = a.get("expenditure")
    if sp:
        sheet("Spend", sp["cost_to_serve_by_channel"],
              ["channel", "cogs_eur", "returns_cost_eur", "discount_leakage_eur",
               "cost_to_serve_eur", "orders"])
        s = wb["Spend"]
        s.append([])
        s.append(["Total COGS", sp["total_cogs_eur"]])
        s.append(["Discount leakage EUR", sp["discount_leakage_eur"]])
        s.append(["Discount leakage %", sp["discount_leakage_pct"]])
        s.append(["Returns cost total", sp["returns_cost_total_eur"]])
        dd = sp.get("leakage_drilldown")
        if dd:
            sheet("Leakage drill-down", dd["by_sales_rep"],
                  ["sales_rep", "region", "leakage_eur", "excess_eur", "within_policy_eur",
                   "leakage_pct_of_revenue", "orders", "orders_above_policy_pct",
                   "median_discount_pct", "p90_discount_pct"])
            s = wb["Leakage drill-down"]
            s.append([])
            s.append(["By region"])
            s.append(["region", "leakage_eur", "excess_eur", "leakage_pct_of_revenue",
                      "orders_above_policy_pct"])
            for r in dd["by_region"]:
                s.append([r["region"], r["leakage_eur"], r["excess_eur"],
                          r["leakage_pct_of_revenue"], r["orders_above_policy_pct"]])
            s.append([])
            s.append(["Policy threshold %", dd["policy_discount_pct"]])
            s.append(["Note", dd["policy_note"]])
    ex = a.get("kpi_alerts")
    if ex:
        rows_x = [{**al, "favorable": ("" if al["favorable"] is None
                                       else ("favourable" if al["favorable"]
                                             else "unfavourable"))}
                  for al in ex["alerts"]]
        sheet("KPI exceptions", rows_x,
              ["label", "month", "value", "center", "lower_limit", "upper_limit",
               "modified_z", "direction", "favorable", "severity"])
        s = wb["KPI exceptions"]
        s.append([])
        s.append(["Control limits (median +/- k*MAD/0.6745), k =", ex["k_sigma"]])
        s.append(["KPI", "centre", "lower_limit", "upper_limit", "months_flagged"])
        for chart in ex["series"].values():
            s.append([chart["label"], chart["center"], chart["lower_limit"],
                      chart["upper_limit"], chart["n_flagged"]])
        s.append([])
        s.append(["Method", ex["method"]])
        s.append(["Note", "Volume series (revenue, orders) excluded from level-anomaly "
                          "detection on 24 points; see forecast CV + revenue bridge. "
                          "Synthetic data."])
    wb.save(out / "kpi_workbook.xlsx")
    return "kpi_workbook.xlsx"


def _write_chart(a: dict, out: Path) -> str | None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    rf = a["revenue_forecast"]
    hist = rf["history"]
    fc = rf["forecast"]
    fig, ax = plt.subplots(figsize=(10, 4))
    xh = list(range(len(hist)))
    xf = list(range(len(hist) - 1, len(hist) + len(fc)))
    ax.plot(xh, hist, color="#2f6bff", label="actual revenue")
    ax.plot(xf, [hist[-1]] + fc, color="#ea4b71", ls="--", marker="o",
            label=f"forecast ({rf['winner']}, CV-MASE {rf['cv_mase']:.2f})")
    band = max(hist) * 0.06
    ax.fill_between(xf, [hist[-1]] + [v - band for v in fc],
                    [hist[-1]] + [v + band for v in fc], color="#ea4b71", alpha=0.15)
    ax.set_title("Monthly revenue — 24 months actual + 3-month forecast")
    ax.set_xlabel("month index")
    ax.set_ylabel("EUR")
    ax.legend(fontsize=8)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(out / "forecast.png", dpi=140)
    plt.close(fig)
    return "forecast.png"


def _svg_esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def render_leakage_svg(dd: dict[str, Any]) -> str:
    """The discount-leakage waterfall as a self-contained, offline SVG string —
    no server, no CDN, no external asset references, pure stdlib.

    The bar geometry and every on-screen number come from `waterfall_steps(dd)`,
    the same model the PNG uses and the tests assert against, so the figure is a
    faithful picture of the computed leakage — never a cosmetic one. Value labels
    are the exact euro amounts (to the cent) so the rendering can be reconciled to
    source in a test.
    """
    steps = waterfall_steps(dd)
    colors = ["#8b8f99", "#2a78d6", "#e34948", "#8b8f99"]   # totals, within, excess, totals
    W, H = 920, 340
    m = {"t": 46, "r": 24, "b": 58, "l": 92}
    iw, ih = W - m["l"] - m["r"], H - m["t"] - m["b"]
    top = dd["gross_list_value_eur"] * 1.06 or 1.0

    def y(v: float) -> float:
        return m["t"] + ih - v / top * ih

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" '
        f'role="img" aria-label="Discount-leakage waterfall from gross list value '
        f'through within-policy and excess discounts to net revenue" '
        f'font-family="system-ui, -apple-system, Segoe UI, Roboto, sans-serif">',
        f'<rect x="0" y="0" width="{W}" height="{H}" fill="#fcfcfb"/>',
        f'<text x="{m["l"]}" y="26" font-size="15" font-weight="700" fill="#0b0b0b">'
        f'Discount-leakage waterfall &#8212; &#8364;{dd["total_leakage_eur"]:,.2f} '
        f'given away vs list</text>',
    ]
    # horizontal gridlines + y-axis ticks (euro, millions for the axis only)
    for t in range(5):
        val = top * t / 4
        yy = y(val)
        parts.append(f'<line x1="{m["l"]}" y1="{yy:.1f}" x2="{W - m["r"]}" y2="{yy:.1f}" '
                     f'stroke="#e1e0d9" stroke-width="1"/>')
        parts.append(f'<text x="{m["l"] - 8}" y="{yy + 3:.1f}" text-anchor="end" '
                     f'font-size="11" fill="#898781">&#8364;{val / 1e6:,.1f}M</text>')

    slot = iw / 4
    bw = 118
    running = [dd["gross_list_value_eur"], dd["gross_list_value_eur"]
               - dd["within_policy_discount_eur"], dd["net_revenue_eur"]]
    for i, (step, color) in enumerate(zip(steps, colors, strict=True)):
        cx = m["l"] + slot * (i + 0.5)
        y_top, y_bot = y(step["bottom"] + step["height"]), y(step["bottom"])
        parts.append(f'<rect x="{cx - bw / 2:.1f}" y="{y_top:.1f}" width="{bw}" '
                     f'height="{max(2.0, y_bot - y_top):.1f}" fill="{color}" rx="4"/>')
        signed = step["value"]
        lab = (f'&#8722;&#8364;{abs(signed):,.2f}' if step["kind"] == "decrease"
               else f'&#8364;{signed:,.2f}')
        parts.append(f'<text x="{cx:.1f}" y="{y_top - 8:.1f}" text-anchor="middle" '
                     f'font-size="12" font-weight="650" fill="#0b0b0b">{lab}</text>')
        for j, line in enumerate(step["label"].split(" ", 1) if " " in step["label"]
                                 else [step["label"]]):
            parts.append(f'<text x="{cx:.1f}" y="{H - 32 + j * 14}" text-anchor="middle" '
                         f'font-size="11" fill="#52514e">{_svg_esc(line)}</text>')
    # dotted connectors at the running totals
    for i, lvl in enumerate(running):
        x1 = m["l"] + slot * (i + 0.5) + bw / 2
        x2 = m["l"] + slot * (i + 1.5) - bw / 2
        parts.append(f'<line x1="{x1:.1f}" y1="{y(lvl):.1f}" x2="{x2:.1f}" y2="{y(lvl):.1f}" '
                     f'stroke="#898781" stroke-width="1" stroke-dasharray="3 4"/>')
    parts.append(
        f'<text x="{m["l"]}" y="{H - 8}" font-size="11" fill="#898781">'
        f'Policy threshold {dd["policy_discount_pct"]:.0f}% of list (assumed) &#183; '
        f'within-policy &#8364;{dd["within_policy_discount_eur"]:,.2f} + excess '
        f'&#8364;{dd["excess_discount_eur"]:,.2f} = &#8364;{dd["total_leakage_eur"]:,.2f} '
        f'leakage</text>')
    parts.append("</svg>")
    return "\n".join(parts) + "\n"


def _write_leakage_svg(a: dict, out: Path) -> str | None:
    """Offline SVG waterfall (stdlib only — always produced, no matplotlib needed)."""
    dd = (a.get("expenditure") or {}).get("leakage_drilldown")
    if not dd:
        return None
    (out / "leakage_waterfall.svg").write_text(render_leakage_svg(dd), encoding="utf-8")
    return "leakage_waterfall.svg"


def _write_leakage_chart(a: dict, out: Path) -> str | None:
    """Discount-leakage waterfall (gross -> within-policy -> excess -> net) plus
    the per-region leakage split. Same decomposition the drill-down reports."""
    dd = (a.get("expenditure") or {}).get("leakage_drilldown")
    if not dd:
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    ink, blue, pink, mute = "#1a1f2b", "#2f6bff", "#ea4b71", "#8b8f99"
    fig, (ax, ax2) = plt.subplots(1, 2, figsize=(12, 4.4), width_ratios=[3, 2])

    within = dd["within_policy_discount_eur"] / 1e6
    gross = dd["gross_list_value_eur"] / 1e6
    net = dd["net_revenue_eur"] / 1e6
    labels = ["Gross\nlist value", "Within-policy\ndiscounts", "Excess\ndiscounts", "Net\nrevenue"]
    # bars come straight from the shared waterfall model — the PNG cannot diverge
    # from the SVG or the tested numbers. Totals sit on the baseline; the two
    # discount cuts float between the running totals.
    colors = [mute, blue, pink, mute]
    for i, (step, color) in enumerate(zip(waterfall_steps(dd), colors, strict=True)):
        bottom, height = step["bottom"] / 1e6, step["height"] / 1e6
        ax.bar(i, height, bottom=bottom, color=color, width=0.62,
               edgecolor="white", linewidth=1.5)
        val = step["value"] / 1e6
        ax.annotate(f"{val:+,.2f}M" if step["kind"] == "decrease" else f"{val:,.2f}M",
                    (i, bottom + height), textcoords="offset points", xytext=(0, 5),
                    ha="center", fontsize=9.5, color=ink)
    # connectors between running totals
    for i, lvl in enumerate([gross, gross - within, net]):
        ax.plot([i + 0.31, i + 1 - 0.31], [lvl, lvl], color=mute, lw=1, ls=":")
    ax.set_xticks(range(4))
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel("EUR (millions)")
    ax.set_title(f"Discount leakage {dd['total_leakage_eur'] / 1e6:,.2f}M EUR = "
                 f"both discount cuts vs list", fontsize=11)
    ax.grid(axis="y", alpha=0.3)
    ax.set_axisbelow(True)

    regs = dd["by_region"]
    names = [r["region"] for r in regs][::-1]
    wp = [r["within_policy_eur"] / 1e6 for r in regs][::-1]
    ex = [r["excess_eur"] / 1e6 for r in regs][::-1]
    ax2.barh(names, wp, color=blue, label="within policy", edgecolor="white", linewidth=1)
    ax2.barh(names, ex, left=wp, color=pink, label="excess", edgecolor="white", linewidth=1)
    for i, r in enumerate(regs[::-1]):
        ax2.text(wp[i] + ex[i], i, f" {r['leakage_eur'] / 1e6:,.2f}M",
                 va="center", fontsize=9, color=ink)
    ax2.set_xlim(0, max(w + e for w, e in zip(wp, ex, strict=False)) * 1.22)
    ax2.set_xlabel("EUR (millions)")
    ax2.set_title(f"Leakage by region (policy = {dd['policy_discount_pct']:.0f}% of list, "
                  f"assumed)", fontsize=11)
    ax2.legend(fontsize=8, frameon=False, loc="lower right")
    ax2.grid(axis="x", alpha=0.3)
    ax2.set_axisbelow(True)
    fig.tight_layout()
    fig.savefig(out / "leakage_waterfall.png", dpi=140)
    plt.close(fig)
    return "leakage_waterfall.png"


# --------------------------------------------------------------------------- #
# Revenue price/volume/mix bridge — CSV + offline SVG waterfall
# --------------------------------------------------------------------------- #
def _write_pvm_csv(a: dict, out: Path) -> str:
    """Per-category revenue PVM decomposition + a TOTAL row that ties out."""
    rb = a.get("revenue_bridge", {})
    cols = ["product_category", "prior_units", "current_units", "prior_price_eur",
            "current_price_eur", "prior_revenue_eur", "current_revenue_eur",
            "price_effect_eur", "volume_effect_eur", "mix_effect_eur"]
    with (out / "pvm_bridge.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        if rb.get("available"):
            for c in rb["by_category"]:
                w.writerow(c)
            w.writerow({
                "product_category": "TOTAL",
                "prior_units": rb["prior_units"], "current_units": rb["current_units"],
                "prior_revenue_eur": rb["prior_revenue_eur"],
                "current_revenue_eur": rb["current_revenue_eur"],
                "price_effect_eur": rb["price_effect_eur"],
                "volume_effect_eur": rb["volume_effect_eur"],
                "mix_effect_eur": rb["mix_effect_eur"],
            })
    return "pvm_bridge.csv"


def render_pvm_svg(rb: dict[str, Any]) -> str:
    """The revenue price/volume/mix bridge as a self-contained, offline SVG —
    no server, no CDN, no external asset references, pure stdlib.

    Bars and every on-screen euro amount come from `bridge_waterfall_steps(rb)`,
    the same model the tests assert against. Two grey totals (prior → current)
    anchor the walk; the three effects float between the running totals. The
    y-axis is zoomed to the delta band (clearly labelled, since the change is
    small next to the ~€11M base) so the effect bars are actually visible; the
    value labels are the exact euro amounts to the cent, so the picture can be
    reconciled to source in a test.
    """
    steps = bridge_waterfall_steps(rb)
    W, H = 940, 380
    m = {"t": 52, "r": 24, "b": 64, "l": 96}
    iw, ih = W - m["l"] - m["r"], H - m["t"] - m["b"]

    # cumulative running totals -> the visible band (zoomed, not from zero)
    cum, run = [], 0.0
    for s in steps[:-1]:
        run = s["value"] if s["kind"] == "total" else run + s["value"]
        cum.append(run)
    lo, hi = min(cum), max(cum)
    span = (hi - lo) or (abs(hi) or 1.0)
    ymin, ymax = lo - 0.14 * span, hi + 0.20 * span

    def y(v: float) -> float:
        return m["t"] + ih - (v - ymin) / (ymax - ymin) * ih

    green, red, grey = "#2f9e6f", "#e34948", "#8b8f99"
    total = rb["total_change_eur"]
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" role="img" '
        f'aria-label="Revenue price-volume-mix bridge from prior-year revenue through '
        f'price, volume and mix effects to current-year revenue" '
        f'font-family="system-ui, -apple-system, Segoe UI, Roboto, sans-serif">',
        f'<rect x="0" y="0" width="{W}" height="{H}" fill="#fcfcfb"/>',
        f'<text x="{m["l"]}" y="26" font-size="15" font-weight="700" fill="#0b0b0b">'
        f'Revenue bridge (YoY) &#8212; price / volume / mix</text>',
        f'<text x="{m["l"]}" y="44" font-size="12" fill="#52514e">'
        f'Total change {"+" if total >= 0 else "&#8722;"}&#8364;{abs(total):,.2f} '
        f'({rb["total_change_pct"]:+.1f}%) &#183; {_svg_esc(rb["period_prior"])} '
        f'&#8594; {_svg_esc(rb["period_current"])}</text>',
    ]
    # horizontal gridlines + zoomed y-axis ticks (euro, millions, 1 dp -> no cents)
    for t in range(5):
        val = ymin + (ymax - ymin) * t / 4
        yy = y(val)
        parts.append(f'<line x1="{m["l"]}" y1="{yy:.1f}" x2="{W - m["r"]}" y2="{yy:.1f}" '
                     f'stroke="#e1e0d9" stroke-width="1"/>')
        parts.append(f'<text x="{m["l"] - 8}" y="{yy + 3:.1f}" text-anchor="end" '
                     f'font-size="11" fill="#898781">&#8364;{val / 1e6:,.1f}M</text>')

    slot = iw / len(steps)
    bw = 108
    for i, s in enumerate(steps):
        cx = m["l"] + slot * (i + 0.5)
        color = grey if s["kind"] == "total" else (green if s["kind"] == "increase" else red)
        if s["kind"] == "total":
            y_top, y_bot = y(s["value"]), y(ymin)
        else:
            y_top, y_bot = y(s["bottom"] + s["height"]), y(s["bottom"])
        parts.append(f'<rect x="{cx - bw / 2:.1f}" y="{y_top:.1f}" width="{bw}" '
                     f'height="{max(2.0, y_bot - y_top):.1f}" fill="{color}" rx="4"/>')
        v = s["value"]
        if s["kind"] == "total":
            lab = f'&#8364;{v:,.2f}'
        elif s["kind"] == "increase":
            lab = f'+&#8364;{v:,.2f}'
        else:
            lab = f'&#8722;&#8364;{abs(v):,.2f}'
        parts.append(f'<text x="{cx:.1f}" y="{y_top - 8:.1f}" text-anchor="middle" '
                     f'font-size="11.5" font-weight="650" fill="#0b0b0b">{lab}</text>')
        parts.append(f'<text x="{cx:.1f}" y="{H - 26}" text-anchor="middle" '
                     f'font-size="11" fill="#52514e">{_svg_esc(s["label"])}</text>')
    # dotted connectors linking each running total to the next bar
    for i, lvl in enumerate(cum):
        x1 = m["l"] + slot * (i + 0.5) + bw / 2
        x2 = m["l"] + slot * (i + 1.5) - bw / 2
        parts.append(f'<line x1="{x1:.1f}" y1="{y(lvl):.1f}" x2="{x2:.1f}" y2="{y(lvl):.1f}" '
                     f'stroke="#898781" stroke-width="1" stroke-dasharray="3 4"/>')
    parts.append(
        f'<text x="{m["l"]}" y="{H - 8}" font-size="11" fill="#898781">'
        f'Realised price = revenue &#247; units per category &#183; volume = proportional '
        f'growth at the prior blended price &#183; mix = residual reallocation &#183; '
        f'y-axis zoomed to the delta band (synthetic data)</text>')
    parts.append("</svg>")
    return "\n".join(parts) + "\n"


def _unit_suffix(unit: str) -> str:
    return "%" if unit == "%" else ""


def _chart_to_render(ex: dict[str, Any]) -> dict[str, Any]:
    """Pick the KPI control chart worth drawing: the one with the most exceptions,
    else the first monitored KPI (so the deliverable is always produced)."""
    series = ex["series"]
    flagged = [c for c in series.values() if c["n_flagged"] > 0]
    if flagged:
        return max(flagged, key=lambda c: (c["n_flagged"], c["kpi"]))
    return next(iter(series.values()))


def render_control_chart_svg(chart: dict[str, Any], k: float = anomaly.DEFAULT_K) -> str:
    """One KPI's robust control chart as a self-contained, offline SVG string —
    no server, no CDN, no external asset references, pure stdlib.

    The centre line, the two control limits and every monthly point come from the
    `control_chart_steps` model the tests assert against, and the numeric labels
    (centre, limits, each out-of-control value) are the exact figures to two
    decimals, so the picture can be reconciled to source in a test.
    """
    st = anomaly.control_chart_steps(chart)
    pts = st["points"]
    unit = _unit_suffix(st["unit"])
    center, lo, hi = st["center"], st["lower_limit"], st["upper_limit"]
    vals = [p["value"] for p in pts]
    ymin = min([lo, *vals])
    ymax = max([hi, *vals])
    pad = (ymax - ymin) * 0.16 or 1.0
    ymin, ymax = ymin - pad, ymax + pad

    W, H = 940, 360
    m = {"t": 58, "r": 128, "b": 52, "l": 66}
    iw, ih = W - m["l"] - m["r"], H - m["t"] - m["b"]
    n = len(pts)

    def x(i: int) -> float:
        return m["l"] + (iw * i / (n - 1) if n > 1 else iw / 2)

    def y(v: float) -> float:
        return m["t"] + ih - (v - ymin) / (ymax - ymin) * ih

    grey, ink, red, blue, mute = "#8b8f99", "#0b0b0b", "#e34948", "#2f6bff", "#898781"
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" role="img" '
        f'aria-label="{_svg_esc(st["label"])} statistical control chart with robust '
        f'centre line and control limits, out-of-control months highlighted" '
        f'font-family="system-ui, -apple-system, Segoe UI, Roboto, sans-serif">',
        f'<rect x="0" y="0" width="{W}" height="{H}" fill="#fcfcfb"/>',
        f'<text x="{m["l"]}" y="26" font-size="15" font-weight="700" fill="{ink}">'
        f'{_svg_esc(st["label"])} &#8212; statistical control monitor</text>',
        f'<text x="{m["l"]}" y="44" font-size="12" fill="#52514e">'
        f'Robust control limits: median &#177; {k:.1f}&#183;MAD/0.6745 '
        f'(Iglewicz&#8211;Hoaglin) &#183; {n} months</text>',
    ]
    # y gridlines + axis ticks
    for t in range(5):
        val = ymin + (ymax - ymin) * t / 4
        yy = y(val)
        parts.append(f'<line x1="{m["l"]}" y1="{yy:.1f}" x2="{m["l"] + iw}" y2="{yy:.1f}" '
                     f'stroke="#e1e0d9" stroke-width="1"/>')
        parts.append(f'<text x="{m["l"] - 8}" y="{yy + 3:.1f}" text-anchor="end" '
                     f'font-size="11" fill="{mute}">{val:,.1f}{unit}</text>')

    # control-limit band + centre line, each labelled with its exact value
    parts.append(f'<rect x="{m["l"]}" y="{y(hi):.1f}" width="{iw}" '
                 f'height="{max(1.0, y(lo) - y(hi)):.1f}" fill="#2f6bff" opacity="0.05"/>')
    for val, lab, col, dash in ((hi, f'UCL {hi:,.2f}{unit}', red, "5 4"),
                                (center, f'&#956; {center:,.2f}{unit}', grey, ""),
                                (lo, f'LCL {lo:,.2f}{unit}', red, "5 4")):
        da = f' stroke-dasharray="{dash}"' if dash else ""
        parts.append(f'<line x1="{m["l"]}" y1="{y(val):.1f}" x2="{m["l"] + iw}" '
                     f'y2="{y(val):.1f}" stroke="{col}" stroke-width="1.4"{da}/>')
        parts.append(f'<text x="{m["l"] + iw + 8}" y="{y(val) + 3:.1f}" font-size="11" '
                     f'font-weight="600" fill="{col}">{lab}</text>')

    # the series polyline
    poly = " ".join(f"{x(i):.1f},{y(p['value']):.1f}" for i, p in enumerate(pts))
    parts.append(f'<polyline points="{poly}" fill="none" stroke="{blue}" stroke-width="1.6"/>')

    # month points; out-of-control ones highlighted with their exact value
    for i, p in enumerate(pts):
        cx, cy = x(i), y(p["value"])
        if p["flagged"]:
            parts.append(f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="5" fill="{red}"/>')
            above = p["value"] >= center
            ly = cy - 9 if above else cy + 17
            parts.append(f'<text x="{cx:.1f}" y="{ly:.1f}" text-anchor="middle" '
                         f'font-size="10.5" font-weight="650" fill="{red}">'
                         f'{p["value"]:,.2f}{unit}</text>')
        else:
            parts.append(f'<circle cx="{cx:.1f}" cy="{cy:.1f}" r="2.6" fill="{blue}"/>')

    # sparse month labels along the x-axis (every third)
    for i in range(0, n, 3):
        parts.append(f'<text x="{x(i):.1f}" y="{H - 30}" text-anchor="middle" '
                     f'font-size="9.5" fill="{mute}">{_svg_esc(pts[i]["month"])}</text>')
    flagged = sum(1 for p in pts if p["flagged"])
    parts.append(
        f'<text x="{m["l"]}" y="{H - 8}" font-size="11" fill="{mute}">'
        f'{flagged} of {n} months out of control (red) &#183; centre = median, '
        f'limits robust to outliers &#183; synthetic data</text>')
    parts.append("</svg>")
    return "\n".join(parts) + "\n"


def _write_control_chart_svg(a: dict, out: Path) -> str | None:
    """Offline SVG KPI control chart (stdlib only — always produced)."""
    ex = a.get("kpi_alerts")
    if not ex or not ex.get("series"):
        return None
    chart = _chart_to_render(ex)
    (out / "kpi_control_chart.svg").write_text(
        render_control_chart_svg(chart, ex["k_sigma"]), encoding="utf-8")
    return "kpi_control_chart.svg"


def _write_control_chart_chart(a: dict, out: Path) -> str | None:
    """Matplotlib PNG of the same KPI control chart (optional — same model as the
    SVG, so the two cannot diverge). Skipped if matplotlib is absent."""
    ex = a.get("kpi_alerts")
    if not ex or not ex.get("series"):
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    chart = _chart_to_render(ex)
    st = anomaly.control_chart_steps(chart)
    pts = st["points"]
    unit = _unit_suffix(st["unit"])
    xs = list(range(len(pts)))
    ys = [p["value"] for p in pts]
    red, blue, grey = "#e34948", "#2f6bff", "#8b8f99"
    fig, ax = plt.subplots(figsize=(11, 4.4))
    ax.axhspan(st["lower_limit"], st["upper_limit"], color=blue, alpha=0.05)
    ax.axhline(st["center"], color=grey, lw=1.4, label=f"centre {st['center']:,.2f}{unit}")
    ax.axhline(st["upper_limit"], color=red, lw=1.3, ls="--",
               label=f"UCL {st['upper_limit']:,.2f}{unit}")
    ax.axhline(st["lower_limit"], color=red, lw=1.3, ls="--",
               label=f"LCL {st['lower_limit']:,.2f}{unit}")
    ax.plot(xs, ys, color=blue, lw=1.6, zorder=2)
    flag_x = [i for i, p in enumerate(pts) if p["flagged"]]
    flag_y = [pts[i]["value"] for i in flag_x]
    ax.scatter([i for i in xs if i not in flag_x],
               [ys[i] for i in xs if i not in flag_x], s=16, color=blue, zorder=3)
    ax.scatter(flag_x, flag_y, s=54, color=red, zorder=4, label="out of control")
    for i, v in zip(flag_x, flag_y, strict=True):
        ax.annotate(f"{v:,.2f}{unit}", (i, v), textcoords="offset points",
                    xytext=(0, -14), ha="center", fontsize=8.5, color=red)
    ax.set_xticks(xs[::3])
    ax.set_xticklabels([pts[i]["month"] for i in xs[::3]], fontsize=8, rotation=45,
                       ha="right")
    ax.set_ylabel(st["label"])
    ax.set_title(f"{st['label']} — robust control chart "
                 f"({sum(1 for p in pts if p['flagged'])} of {len(pts)} months out of control)",
                 fontsize=11)
    ax.legend(fontsize=8, frameon=False, loc="lower left", ncol=2)
    ax.grid(axis="y", alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    fig.savefig(out / "kpi_control_chart.png", dpi=140)
    plt.close(fig)
    return "kpi_control_chart.png"


def _write_pvm_svg(a: dict, out: Path) -> str | None:
    """Offline SVG revenue-bridge waterfall (stdlib only — always produced)."""
    rb = a.get("revenue_bridge", {})
    if not rb.get("available"):
        return None
    (out / "pvm_waterfall.svg").write_text(render_pvm_svg(rb), encoding="utf-8")
    return "pvm_waterfall.svg"


def _write_pvm_chart(a: dict, out: Path) -> str | None:
    """Matplotlib PNG of the revenue-bridge waterfall (optional — same model as
    the SVG, so the two cannot diverge). Skipped if matplotlib is absent."""
    rb = a.get("revenue_bridge", {})
    if not rb.get("available"):
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return None
    ink, green, red, grey = "#1a1f2b", "#2f9e6f", "#2f6bff", "#8b8f99"
    steps = bridge_waterfall_steps(rb)
    fig, ax = plt.subplots(figsize=(11, 4.6))
    cum, run = [], 0.0
    for s in steps[:-1]:
        run = s["value"] if s["kind"] == "total" else run + s["value"]
        cum.append(run)
    lo, hi = min(cum), max(cum)
    span = (hi - lo) or 1.0
    ymin, ymax = (lo - 0.14 * span) / 1e6, (hi + 0.20 * span) / 1e6
    for i, s in enumerate(steps):
        color = grey if s["kind"] == "total" else (green if s["kind"] == "increase" else red)
        if s["kind"] == "total":
            bottom, height = ymin, s["value"] / 1e6 - ymin
        else:
            bottom, height = s["bottom"] / 1e6, s["height"] / 1e6
        ax.bar(i, height, bottom=bottom, color=color, width=0.6,
               edgecolor="white", linewidth=1.4)
        v = s["value"] / 1e6
        lab = f"{v:,.2f}M" if s["kind"] == "total" else f"{v:+,.2f}M"
        ax.annotate(lab, (i, bottom + height), textcoords="offset points",
                    xytext=(0, 5), ha="center", fontsize=9.5, color=ink)
    for i, lvl in enumerate(cum):
        ax.plot([i + 0.3, i + 1 - 0.3], [lvl / 1e6, lvl / 1e6], color=grey, lw=1, ls=":")
    ax.set_ylim(ymin, ymax)
    ax.set_xticks(range(len(steps)))
    ax.set_xticklabels([s["label"] for s in steps], fontsize=9)
    ax.set_ylabel("EUR (millions)")
    ax.set_title(f"Revenue bridge {rb['total_change_eur'] / 1e6:+,.2f}M EUR YoY "
                 f"= price {rb['price_effect_eur'] / 1e3:+,.0f}k + volume "
                 f"{rb['volume_effect_eur'] / 1e6:+,.2f}M + mix "
                 f"{rb['mix_effect_eur'] / 1e6:+,.2f}M", fontsize=10.5)
    ax.grid(axis="y", alpha=0.3)
    ax.set_axisbelow(True)
    fig.tight_layout()
    fig.savefig(out / "pvm_waterfall.png", dpi=140)
    plt.close(fig)
    return "pvm_waterfall.png"
